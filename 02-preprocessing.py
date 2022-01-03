from openpyxl import load_workbook
from openpyxl import Workbook
import re

# 전처리
load_wb = load_workbook("C:/Users/user/OneDrive/바탕 화면/whdrkd/01-카카오.xlsx", data_only=True)
load_ws = load_wb['Sheet']

lst_excel = [] # 저장   
num = 1 # 행 수

while load_ws.cell(num, 2).value != None:
    date = load_ws.cell(num, 1).value
    text = load_ws.cell(num, 2).value
    num += 1
    
    rep = text.replace('.',' ')
    rep = rep.replace('\"','')
    rep = rep.replace('\'','')
    rep = rep.replace(',',' ')
    rep = rep.replace('-',' ')
    rep = rep.replace('·',' ')
    rep = rep.replace('・',' ')
    rep = rep.replace('?',' ')
    rep = rep.replace('…',' ')
    rep = rep.replace('“','')
    rep = rep.replace('”','')
    rep = rep.replace('‘','')
    rep = rep.replace('’','')
    rep = rep.replace('①',' ')
    rep = rep.replace('②',' ')
    rep = rep.replace('③',' ')
    rep = rep.replace('⑤',' ')
    rep = rep.replace('④',' ')
    rep = rep.replace('/',' ')
    rep = rep.replace('•',' ')
    rep = rep.replace('↑',' ↑')
    rep = rep.replace('↓',' ↓')
    rep = rep.replace('+',' ')
    rep = rep.replace('|',' ')
    rep = rep.replace('‧',' ')
    rep = rep.replace('!',' ')
    rep = rep.replace('`',' ')
  
    p = re.compile("(에도|에서도|에는|에게도|에서는|에게|에서|에|까지|(?P<의>[^혐합논건])의|(?P<도>[^속매])도|대로|\
        (?P<으로>[^앞])으로|(?P<로>[^애구])로)\s")
    rep = p.sub("\g<의>\g<도>\g<으로>\g<로> ",rep)

    rep = re.sub("\[(.)+\]"," ",rep)
    rep = re.sub("\((.)+\)"," ",rep)
    rep = re.sub("\<(.)+\>"," ",rep)
    rep = rep.replace('[',' ')
    rep = rep.replace('(',' ')
    rep = rep.replace('<',' ')
    
    lst_excel.append((date,rep))

write_wb = Workbook()
write_ws = write_wb.active
n = 1 # 행 번호

for d, l in lst_excel:
    write_ws.cell(n,1,d)
    write_ws.cell(n,2,l)
    n += 1
    
write_wb.save("C:/Users/user/OneDrive/바탕 화면/whdrkd/02-전처리-카카오.xlsx")
