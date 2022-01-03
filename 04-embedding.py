from gensim.models import Word2Vec
from openpyxl import load_workbook
from openpyxl import Workbook

# 워드 임베딩
load_wb = load_workbook("C:/Users/user/OneDrive/바탕 화면/whdrkd/02-전처리-카카오.xlsx", data_only=True)
load_ws = load_wb['Sheet']

num = 1
lst = []

while load_ws.cell(num, 1).value != None:
    load = load_ws.cell(num, 2).value
    num += 1
    if load == None: continue
    lst.append(load.split())

model = Word2Vec(sentences=lst, vector_size=300, window=5, min_count=5, workers=4, sg=0)

excel_pos='''출시
상승
최고
강세
증가
기대
인기
강보합
주목
혁신
하락
우려
약세
논란
중단
급락
축소
약보합
비상
순매도
부진'''

word = excel_pos.split('\n')

write_wb = Workbook()
write_ws = write_wb.active
n=1

for i in word:
    model_result = model.wv.most_similar(i) # 임베딩 진행
    write_ws.cell(n,1,i) 
    n+=1
    for j in model_result:
        write_ws.cell(n,1,j[0])
        write_ws.cell(n,2,j[1])
        n+=1
    
write_wb.save("C:/Users/user/OneDrive/바탕 화면/whdrkd/04-임베딩-카카오.xlsx")
