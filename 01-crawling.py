import requests
from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import Workbook

# 크롤링
def craw():
    num = 1 # 페이지 번호
    global num_excel # 엑셀 행 번호

    for i in range(20): # 페이지 수 (한 페이지 당 10건)
        url = 'https://search.naver.com/search.naver?where=news&sm=tab_pge&query=' + item + '&sort=0&photo=0&field=0&pd=3&ds=' + from_date + '&de=' + to_date + '&cluster_rank=65&mynews=0&office_type=0&office_section_code=0&news_office_checked=&nso=so:r,p:from' + from_date.strip('.') + 'to' + to_date.strip('.') + ',a:all&is_sug_officeid=1&start=' + str(num)
        response = requests.get(url)

        if response.status_code == 200:
            html = response.text
            soup = BeautifulSoup(html, 'html.parser')
            ul = soup.select_one('ul.list_news')
            titles = ul.select('li > div > div > a') # api
            for title in titles: # 엑셀 기록
                write_ws.cell(num_excel,1,from_date)
                write_ws.cell(num_excel,2,title.get_text())
                num_excel += 1
        else : print(response.status_code)

        num += 10 # 네이버 뉴스 페이지 url은 +10
        
def ddate(month,date): # url 날짜
    global from_date
    global to_date
    
    if month<10 and date<10:
        from_date = '2021.0'+str(month)+'.0'+str(date)
        to_date = '2021.0'+str(month)+'.0'+str(date)
    elif month<10 and date>=10:
        from_date = '2021.0'+str(month)+'.'+str(date)
        to_date = '2021.0'+str(month)+'.'+str(date)
    elif month>=10 and date<10:
        from_date = '2021.'+str(month)+'.0'+str(date)
        to_date = '2021.'+str(month)+'.0'+str(date)
    else: 
        from_date = '2021.'+str(month)+'.'+str(date)
        to_date = '2021.'+str(month)+'.'+str(date)

item = '카카오' # 대상
num_excel = 1

write_wb = Workbook() # 엑셀
write_ws = write_wb.active

for m in range(1,13): # 월
    for d in range(1,32): # 일
        if m==2 and d>28: break
        if m in [2,4,6,9,11] and d==31: break
        ddate(m,d)
        craw()
        
write_wb.save("C:/Users/user/OneDrive/바탕 화면/whdrkd/01-카카오.xlsx")