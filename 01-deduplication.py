from openpyxl import load_workbook
from openpyxl import Workbook

# 중복되는 뉴스 제거
load_wb = load_workbook("C:/Users/user/OneDrive/바탕 화면/whdrkd/01-카카오.xlsx", data_only=True)
load_ws = load_wb['Sheet']

lst_excel = [] # 엑셀에 기록할 요소 저장
num = 1 # 엑셀 행 번호
preText = " " # 이전 뉴스 내용

while load_ws.cell(num, 2).value != None: # 엑셀 끝나면 stop
    d = load_ws.cell(num, 1).value # 날짜
    text = load_ws.cell(num, 2).value # 뉴스 내용
    num += 1
    
    if preText != text: # 이전 뉴스와 중복이 아니면 저장
        lst_excel.append((d,text)) 
    
    preText = text # 지금 뉴스가 이전 뉴스 됨
    
write_wb = Workbook()
write_ws = write_wb.active
n=1 # 행 개수

for d, l in lst_excel:
    write_ws.cell(n,1,d)
    write_ws.cell(n,2,l)
    n+=1

write_wb.save("C:/Users/user/OneDrive/바탕 화면/whdrkd/01-카카오-deduplication.xlsx")
