from openpyxl import load_workbook
from openpyxl import Workbook

# 빈도목록
load_wb = load_workbook("C:/Users/user/OneDrive/바탕 화면/whdrkd/02-전처리-카카오.xlsx", data_only=True)
load_ws = load_wb['Sheet']

lst_excel = []
word_count = {} # 단어 개수 세기
num = 1

while load_ws.cell(num, 1).value != None:
    text = load_ws.cell(num, 2).value
    num += 1
    spl = text.split()
    
    for i in spl:
        if i in word_count: word_count[i] += 1
        else: word_count[i] = 1
       
sortWC = dict(sorted(word_count.items(), key=lambda x: x[1], reverse=True)) # key 기준 내림차순 정렬
 
write_wb = Workbook()
write_ws = write_wb.active
n=1
for k,v in sortWC.items():
    write_ws.cell(n,1,k)
    write_ws.cell(n,2,v)
    n+=1
    
write_wb.save("C:/Users/user/OneDrive/바탕 화면/whdrkd/03-빈도-카카오.xlsx")