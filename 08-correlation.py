import matplotlib.pyplot as plt
from openpyxl import load_workbook
import scipy.stats as stats
import pandas as pd
import numpy as np

load_wb = load_workbook("C:/Users/user/OneDrive/바탕 화면/whdrkd/06-polarity-카카오.xlsx", data_only=True)
load_ws = load_wb['Sheet']

load_wb2 = load_workbook("C:/Users/user/OneDrive/바탕 화면/whdrkd/07-일별주가-카카오.xlsx", data_only=True)
load_ws2 = load_wb2['Sheet1']

date = []
val = []
for i in range(2,250):
    load_date = load_ws2.cell(i, 1).value # 일자
    date.append((int(load_date[5]+load_date[6]),int(load_date[8]+load_date[9])))

    load_val = float(load_ws2.cell(i, 4).value) # 주가 등락
    val.insert(0,load_val) # 시간 순으로 정렬

lstg = []
for i in range(2,367):
    # 주식 개장일 데이터 선별
    if (int(load_ws.cell(i, 1).value),int(load_ws.cell(i, 2).value)) in date:
        data = float(load_ws.cell(i, 5).value)
        lstg.append(data)

# 시각화
a = plt.subplot()
l1 = a.plot(lstg,'b', label = 'polarity')
b = a.twinx()
l2 = b.plot(val,'r',label = "stock")
a.set_xlabel('date')
a.set_ylabel("polarity")
b.set_ylabel("stock")
lines = l1 + l2
labels = [l.get_label() for l in lines]
a.legend(lines, labels, loc='upper right')
plt.show()

plt.scatter(lstg,val)
plt.xlabel("polarity"); plt.ylabel("stock")
plt.show()

print(np.corrcoef(lstg,val)[0,1])
print(stats.pearsonr(lstg,val))