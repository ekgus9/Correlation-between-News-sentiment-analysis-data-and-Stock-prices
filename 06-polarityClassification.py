from openpyxl import load_workbook
from openpyxl import Workbook

# 데이터 극성 분류
load_wb = load_workbook("C:/Users/user/OneDrive/바탕 화면/whdrkd/02-전처리-카카오.xlsx", data_only=True)
load_ws = load_wb['Sheet']
load_wb2 = load_workbook("C:/Users/user/OneDrive/바탕 화면/whdrkd/05-감성사전.xlsx", data_only=True)
load_ws2 = load_wb2['Sheet1']

pos_word = []
neg_word = []

for i in range(1,81): # 감성 사전 불러옴
    load_pos = load_ws2.cell(i, 1).value
    if load_pos != None: pos_word.append(load_pos)
        
    load_neg = load_ws2.cell(i, 2).value
    if load_neg != None: neg_word.append(load_neg)

lst = []
num = 1
pre_month = 0
pre_date = 0

rate = 0
pos = 0
neg = 0

while load_ws.cell(num, 1).value != None:
    load_date = load_ws.cell(num, 1).value # 뉴스 데이터
    load_news = load_ws.cell(num, 2).value
    num += 1
    if load_news == None: continue # 뉴스 내용 없으면 skip

    month = int(load_date[5] + load_date[6])
    date = int(load_date[8] + load_date[9])

    if month != pre_month or date != pre_date: # 같은 일자별 분류
        if pos + neg != 0: rate = pos / (pos + neg) # 긍정 비율 계산
        lst.append((pre_month,pre_date,pos,neg,rate))
        pos = 0; neg = 0
    
    # 감성 점수 계산
    if sum([True for p in pos_word if p in load_news]) - sum([True for n in neg_word if n in load_news])> 0: pos += 1
    elif sum([True for p in pos_word if p in load_news]) - sum([True for n in neg_word if n in load_news])< 0: neg += 1

    pre_month = month
    pre_date = date

rate = pos / (pos + neg)    
lst.append((month,date,pos,neg,rate)) # 마지막 값 추가

write_wb = Workbook()
write_ws = write_wb.active

write_ws.cell(1,1,"month")
write_ws.cell(1,2,"date")
write_ws.cell(1,3,"pos")
write_ws.cell(1,4,"neg")
write_ws.cell(1,5,"rate")
n=2

for i in lst: 
    write_ws.cell(n,1,i[0])
    write_ws.cell(n,2,i[1])
    write_ws.cell(n,3,i[2])
    write_ws.cell(n,4,i[3])
    write_ws.cell(n,5,i[4])
    n+=1

write_wb.save("C:/Users/user/OneDrive/바탕 화면/whdrkd/06-polarity-카카오.xlsx")